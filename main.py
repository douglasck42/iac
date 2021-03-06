#!/usr/bin/env python3

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side 
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
import aci_lib
import getpass
import re, os, sys
import subprocess
import time

# Global Variables
excel_workbook = None
home = Path.home()

Access_regex = re.compile('(add_apg|vlan_pool)')
Admin_regex = re.compile('(backup|radius|tacacs|realm|web_security)')
Contracts_regex = re.compile('(add_contract|add_subject|add_filter)')
DHCP_regex = re.compile('(add_vrf|ctx_common)')
Fabric_regex = re.compile('(bgp_(as|rr)|dns|dns_mgmt|domain|ntp|smartcallhome|snmp_(client|comm|info|trap|user)|syslog_(dg|rmt))')
Inventory_regex = re.compile('(apic_inb|inb_subnet|switch|vpc_pair)')
L3Out_regex = re.compile('(add_l3out|node_intf|node_prof)')
networks_regex = re.compile('(add_net)')
Sites_regex = re.compile('(site_id|grp_id)')
Tenant_regex = re.compile('(add_tenant)')
VRF_regex = re.compile('(add_vrf|ctx_common)')
VMM_regex = re.compile('(add_vrf|ctx_common)')

def apply_aci_terraform(folders):

    print(f'\n-----------------------------------------------------------------------------\n')
    print(f'  Found the Followng Folders with uncommitted changes:\n')
    for folder in folders:
        print(f'  - {folder}')
    print(f'\n  Beginning Terraform Plan and Apply in each folder.')
    print(f'\n-----------------------------------------------------------------------------\n')

    time.sleep(7)

    response_p = ''
    response_a = ''
    for folder in folders:
        path = './%s' % (folder)
        p = subprocess.Popen(['terraform', 'init', '-plugin-dir=../../../terraform-plugins/providers/'], cwd=path)
        p.wait()
        p = subprocess.Popen(['terraform', 'plan', '-out=main.plan'], cwd=path)
        p.wait()
        while True:
            print(f'\n-----------------------------------------------------------------------------\n')
            print(f'  Terraform Plan Complete.  Please Review the Plan and confirm if you want')
            print(f'  to move forward.  "A" to Apply the Plan. "S" to Skip.  "Q" to Quit.')
            print(f'\n-----------------------------------------------------------------------------\n')
            response_p = input('  Please Enter ["A", "S" or "Q"]: ')
            if re.search('^(A|S)$', response_p):
                break
            elif response_p == 'Q':
                exit()
            else:
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'  A Valid Response is either "A", "S" or "Q"...')
                print(f'\n-----------------------------------------------------------------------------\n')

        if response_p == 'A':
            p = subprocess.Popen(['terraform', 'apply', '-parallelism=1', 'main.plan'], cwd=path)
            p.wait()

        while True:
            if response_p == 'A':
                response_p = ''
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'  Terraform Apply Complete.  Please Review for any errors and confirm if you')
                print(f'  want to move forward.  "M" to Move to the Next Section. "Q" to Quit..')
                print(f'\n-----------------------------------------------------------------------------\n')
                response_a = input('  Please Enter ["M" or "Q"]: ')
            elif response_p == 'S':
                break
            if response_a == 'M':
                break
            elif response_a == 'Q':
                exit()
            else:
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'  A Valid Response is either "M" or "Q"...')
                print(f'\n-----------------------------------------------------------------------------\n')

def check_git_status():
    random_folders = []
    git_path = './'
    result = subprocess.Popen(['python3', '-m', 'git_status_checker'], stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    while(True):
        # returns None while subprocess is running
        retcode = result.poll()
        line = result.stdout.readline()
        line = line.decode('utf-8')
        if re.search(r'M (.*/).*.tf\n', line):
            folder = re.search(r'M (.*/).*.tf\n', line).group(1)
            if not folder in random_folders:
                random_folders.append(folder)
        if retcode is not None:
            break

    if not random_folders:
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   There were no uncommitted changes in the environment.')
        print(f'   Proceedures Complete!!! Closing Environment and Exiting Script.')
        print(f'\n-----------------------------------------------------------------------------\n')
        exit()

    strict_folders = []
    folder_order = ['Tenant_mgmt', 'Access', 'VLANs', 'Fabric', 'Admin', 'Tenant_common', 'Tenant_infra', 'L3Out']
    for folder in folder_order:
        for fx in random_folders:
            if folder in fx:
                strict_folders.append(fx)
                random_folders.remove(fx)
    for folder in random_folders:
        strict_folders.append(folder)
        random_folders.remove(folder)
    
    return strict_folders

def get_user_pass():
    print(f'\n-----------------------------------------------------------------------------\n')
    print(f'   Beginning Proceedures to Apply Terraform Resources to the environment')
    print(f'\n-----------------------------------------------------------------------------\n')

    user = input('Enter APIC username: ')
    while True:
        try:
            password = getpass.getpass(prompt='Enter APIC password: ')
            break
        except Exception as e:
            print('Something went wrong. Error received: {}'.format(e))

    os.environ['TF_VAR_aciUser'] = '%s' % (user)
    os.environ['TF_VAR_aciPass'] = '%s' % (password)

def process_Access(wb):
    # Evaluate Access Worksheet
    ws = wb['Access']
    aci_lib_ref = 'aci_lib.Access_Policies'
    func_regex = Access_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)
    
    # Evaluate Inventory Worksheet
    ws = wb['Inventory']
    func_regex = Inventory_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

def process_Admin(wb):
    # Evaluate Admin Worksheet
    ws = wb['Admin']
    aci_lib_ref = 'aci_lib.Admin_Policies'
    func_regex = Admin_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)
    
def process_Contracts(wb):
    # Evaluate Fabric Worksheet
    ws = wb['Contracts']
    aci_lib_ref = 'aci_lib.Contracts_Policies'
    func_regex = Contracts_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

def process_Fabric(wb):
    # Evaluate Fabric Worksheet
    ws = wb['Fabric']
    aci_lib_ref = 'aci_lib.Fabric_Policies'
    func_regex = Fabric_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

def process_L3Out(wb):
    # Evaluate Fabric Worksheet
    ws = wb['L3Out']
    aci_lib_ref = 'aci_lib.L3Out_Policies'
    func_regex = L3Out_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

def process_Sites(wb):
    # Evaluate Sites Worksheet
    ws = wb['Sites']
    aci_lib_ref = 'aci_lib.Site_Policies'
    func_regex = Sites_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

def process_Tenants(wb):
    # Evaluate Tenants Worksheet
    ws = wb['Tenants']
    aci_lib_ref = 'aci_lib.Tenant_Policies'
    func_regex = Tenant_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

    # Evaluate VRF Worksheet
    ws = wb['VRF']
    func_regex = VRF_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

    # Evaluate Network Segments Worksheet
    ws = wb['Networks']
    func_regex = networks_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

    # Evaluate DHCP Relay
    # ws = wb['DHCP Relay']
    # func_regex = DHCP_regex
    # read_worksheet(wb, ws, aci_lib_ref, func_regex)
    
def process_VMM(wb):
    # Evaluate Sites Worksheet
    ws = wb['VMM']
    aci_lib_ref = 'aci_lib.VMM_Policies'
    func_regex = VMM_regex
    read_worksheet(wb, ws, aci_lib_ref, func_regex)

def read_worksheet(wb, ws, aci_lib_ref, func_regex):
    rows = ws.max_row
    func_list = aci_lib.findKeys(ws, func_regex)
    class_init = '%s(ws)' % (aci_lib_ref)
    aci_lib.stdout_log(ws, None)
    for func in func_list:
        count = aci_lib.countKeys(ws, func)
        var_dict = aci_lib.findVars(ws, func, rows, count)
        for pos in var_dict:
            row_num = var_dict[pos]['row']
            del var_dict[pos]['row']
            for x in list(var_dict[pos].keys()):
                if var_dict[pos][x] == '':
                    del var_dict[pos][x]
            aci_lib.stdout_log(ws, row_num)
            eval("%s.%s(wb, ws, row_num, **var_dict[pos])" % (class_init, func))

def wb_update(wr_ws, status, i):
    # build green and red style sheets for excel
    bd1 = Side(style="thick", color="8EA9DB")
    bd2 = Side(style="medium", color="8EA9DB")
    wsh1 = NamedStyle(name="wsh1")
    wsh1.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
    wsh1.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
    wsh1.font = Font(bold=True, size=15, color="FFFFFF")
    wsh2 = NamedStyle(name="wsh2")
    wsh2.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
    wsh2.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    wsh2.fill = PatternFill("solid", fgColor="305496")
    wsh2.font = Font(bold=True, size=15, color="FFFFFF")
    green_st = NamedStyle(name="ws_odd")
    green_st.alignment = Alignment(horizontal="center", vertical="center")
    green_st.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    green_st.fill = PatternFill("solid", fgColor="D9E1F2")
    green_st.font = Font(bold=False, size=12, color="44546A")
    red_st = NamedStyle(name="ws_even")
    red_st.alignment = Alignment(horizontal="center", vertical="center")
    red_st.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    red_st.font = Font(bold=False, size=12, color="44546A")
    yellow_st = NamedStyle(name="ws_even")
    yellow_st.alignment = Alignment(horizontal="center", vertical="center")
    yellow_st.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    yellow_st.font = Font(bold=False, size=12, color="44546A")
    # green_st = xlwt.easyxf('pattern: pattern solid;')
    # green_st.pattern.pattern_fore_colour = 3
    # red_st = xlwt.easyxf('pattern: pattern solid;')
    # red_st.pattern.pattern_fore_colour = 2
    # yellow_st = xlwt.easyxf('pattern: pattern solid;')
    # yellow_st.pattern.pattern_fore_colour = 5
    # if stanzas to catch the status code from the request
    # and then input the appropriate information in the workbook
    # this then writes the changes to the doc
    if status == 200:
        wr_ws.write(i, 1, 'Success (200)', green_st)
    if status == 400:
        print("Error 400 - Bad Request - ABORT!")
        print("Probably have a bad URL or payload")
        wr_ws.write(i, 1, 'Bad Request (400)', red_st)
        pass
    if status == 401:
        print("Error 401 - Unauthorized - ABORT!")
        print("Probably have incorrect credentials")
        wr_ws.write(i, 1, 'Unauthorized (401)', red_st)
        pass
    if status == 403:
        print("Error 403 - Forbidden - ABORT!")
        print("Server refuses to handle your request")
        wr_ws.write(i, 1, 'Forbidden (403)', red_st)
        pass
    if status == 404:
        print("Error 404 - Not Found - ABORT!")
        print("Seems like you're trying to POST to a page that doesn't"
              " exist.")
        wr_ws.write(i, 1, 'Not Found (400)', red_st)
        pass
    if status == 666:
        print("Error - Something failed!")
        print("The POST failed, see stdout for the exception.")
        wr_ws.write(i, 1, 'Unkown Failure', yellow_st)
        pass
    if status == 667:
        print("Error - Invalid Input!")
        print("Invalid integer or other input.")
        wr_ws.write(i, 1, 'Unkown Failure', yellow_st)
        pass

def main():
    # Ask user for required Information: ACI_DEPLOY_FILE
    if sys.argv[1:]:
        if os.path.isfile(sys.argv[1]):
            excel_workbook = sys.argv[1]
        else:
            print('\nWorkbook not Found.  Please enter a valid /path/filename for the source you will be using.')
            while True:
                print('Please enter a valid /path/filename for the source you will be using.')
                excel_workbook = input('/Path/Filename: ')
                if os.path.isfile(excel_workbook):
                    print(f'\n-----------------------------------------------------------------------------\n')
                    print(f'   {excel_workbook} exists.  Will Now Check for API Variables...')
                    print(f'\n-----------------------------------------------------------------------------\n')
                    break
                else:
                    print('\nWorkbook not Found.  Please enter a valid /path/filename for the source you will be using.')
    else:
        while True:
            print('Please enter a valid /path/filename for the source you will be using.')
            excel_workbook = input('/Path/Filename: ')
            if os.path.isfile(excel_workbook):
                print(f'\n-----------------------------------------------------------------------------\n')
                print(f'   {excel_workbook} exists.  Will Now Check for API Variables...')
                print(f'\n-----------------------------------------------------------------------------\n')
                break
            else:
                print('\nWorkbook not Found.  Please enter a valid /path/filename for the source you will be using.')

    # Load Workbook
    wb = aci_lib.read_in(excel_workbook)
    
    # Run Proceedures for Worksheets in the Workbook
    process_Sites(wb)

    # Either Run All Remaining Proceedures or Just Specific based on sys.argv[2:]
    if sys.argv[2:]:
        if re.search('access', str(sys.argv[2:])):
            process_Access(wb)
        elif re.search('admin', str(sys.argv[2:])):
            process_Admin(wb)
        elif re.search('contracts', str(sys.argv[2:])):
            process_Contracts(wb)
        elif re.search('fabric', str(sys.argv[2:])):
            process_Fabric(wb)
        elif re.search('l3out', str(sys.argv[2:])):
            process_L3Out(wb)
        elif re.search('tenant', str(sys.argv[2:])):
            process_Tenants(wb)
        elif re.search('vmm', str(sys.argv[2:])):
            process_VMM(wb)
        else:
            process_Fabric(wb)
            process_Access(wb)
            process_Admin(wb)
            process_Contracts(wb)
            process_L3Out(wb)
            process_Tenants(wb)
    else:
        process_Fabric(wb)
        process_Access(wb)
        process_Admin(wb)
        process_Tenants(wb)

    folders = check_git_status()
    get_user_pass()
    apply_aci_terraform(folders)

    print(f'\n-----------------------------------------------------------------------------\n')
    print(f'  Proceedures Complete!!! Closing Environment and Exiting Script.')
    print(f'\n-----------------------------------------------------------------------------\n')
    exit()
    
if __name__ == '__main__':
    main()